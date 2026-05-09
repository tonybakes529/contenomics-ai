"""
Builds tracker.xlsx — YouTube content tracker for Viewtrack.

Run:  python3 build_tracker.py
Output: tracker.xlsx in the same directory.

Design notes:
- No fills, no custom colors, no header bands. Default formatting only.
- Arial throughout (header rows: bold; everything else: regular).
- 50 pre-formatted blank rows on Identity. Sister sheets keyed off Video ID
  (col A) auto-pull the ID from Identity via formula so you only type IDs once.
- Dropdowns implemented as DataValidation list constraints.
- Dashboard + AI Export are formula-driven (INDEX/MATCH + AVERAGEIF).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROWS = 50  # number of pre-formatted data rows on each sheet
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True)
BODY_FONT = Font(name=ARIAL)


# ---------- helpers ---------------------------------------------------------


def write_headers(ws, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def apply_body_font(ws, n_cols: int, n_rows: int = ROWS) -> None:
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).font = BODY_FONT


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dropdown(ws, options: list[str], col_letter: str, n_rows: int = ROWS) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Pick a value from the list"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{n_rows + 1}")


def id_formula(target_row: int) -> str:
    """Pull the Video ID from Identity!A{row} so sister sheets stay aligned."""
    return f"=IF(Identity!A{target_row}=\"\",\"\",Identity!A{target_row})"


# ---------- workbook --------------------------------------------------------

wb = Workbook()


# ===== Sheet 1 — Identity ===================================================

identity = wb.active
identity.title = "Identity"

identity_headers = [
    "Video ID",
    "Date Posted",
    "Final Title",
    "Video URL",
    "Length",
    "Topic",
]
write_headers(identity, identity_headers)

# Auto-format Video ID as LF-001, LF-002, ...  Pre-fill all 50 rows so the ID
# always exists; the sister sheets use these for INDEX/MATCH.
for i in range(ROWS):
    row = i + 2
    identity.cell(row=row, column=1, value=f"LF-{i + 1:03d}")

# Date column: a stable display format.
for r in range(2, ROWS + 2):
    identity.cell(row=r, column=2).number_format = "yyyy-mm-dd"
    # Length is free-form text (e.g., "8:42") — keep as text.
    identity.cell(row=r, column=5).number_format = "@"

apply_body_font(identity, len(identity_headers))
autosize(identity, [10, 13, 42, 36, 10, 22])
identity.freeze_panes = "A2"


# ===== Sheet 2 — Creative ===================================================

creative = wb.create_sheet("Creative")
creative_headers = [
    "Video ID",
    "A/B Title 1",
    "A/B Title 2",
    "A/B Title 3",
    "Winning Title #",
    "Winning Style",
    "Thumbnail URL",
    "Face Y/N",
    "Face Emotion",
    "Word Count",
    "Words Used",
    "Background",
    "Color Palette",
    "Hook Script",
    "Hook Style",
    "Intro Length sec",
    "Time-to-Value sec",
]
write_headers(creative, creative_headers)

# Pull Video ID from Identity.
for i in range(ROWS):
    r = i + 2
    creative.cell(row=r, column=1, value=id_formula(r))

# Dropdowns
add_dropdown(creative, ["1", "2", "3"], "E")
add_dropdown(creative, ["Curiosity", "List", "How-To", "Question", "Shock"], "F")
add_dropdown(creative, ["Y", "N"], "H")
add_dropdown(creative, ["Shock", "Smile", "Serious", "Curious", "None"], "I")
add_dropdown(creative, ["Solid", "Scene", "Gradient", "Photo"], "L")
add_dropdown(creative, ["Bright", "Dark", "Mixed"], "M")
add_dropdown(
    creative,
    ["Question", "Stat", "Story", "Bold Claim", "Pattern Interrupt"],
    "O",
)

apply_body_font(creative, len(creative_headers))
autosize(
    creative,
    [10, 28, 28, 28, 14, 14, 36, 9, 14, 12, 24, 13, 14, 42, 18, 16, 18],
)
creative.freeze_panes = "B2"


# ===== Sheet 3 — Performance ================================================

perf = wb.create_sheet("Performance")
perf_headers = [
    "Video ID",
    "24hr CTR%",
    "7-day CTR%",
    "30-day CTR%",
    "Drop-off Rate%",
    "Drop-off Timestamp",
    "AVD",
    "7-day Views",
    "30-day Views",
]
write_headers(perf, perf_headers)
for i in range(ROWS):
    r = i + 2
    perf.cell(row=r, column=1, value=id_formula(r))

# Number formats: percentages on B-E, plain text timestamp on F, AVD as text
# (e.g., "4:32"), counts on H-I.
for r in range(2, ROWS + 2):
    for col in (2, 3, 4, 5):
        perf.cell(row=r, column=col).number_format = "0.00%"
    perf.cell(row=r, column=6).number_format = "@"
    perf.cell(row=r, column=7).number_format = "@"
    for col in (8, 9):
        perf.cell(row=r, column=col).number_format = "#,##0"

apply_body_font(perf, len(perf_headers))
autosize(perf, [10, 12, 12, 12, 14, 18, 10, 14, 14])
perf.freeze_panes = "B2"


# ===== Sheet 4 — Conversion =================================================

conv = wb.create_sheet("Conversion")
conv_headers = [
    "Video ID",
    "CTAs Used",
    "Click-throughs",
    "Engagement Rate",
    "Calls Booked",
    "Notes",
]
write_headers(conv, conv_headers)
for i in range(ROWS):
    r = i + 2
    conv.cell(row=r, column=1, value=id_formula(r))
for r in range(2, ROWS + 2):
    conv.cell(row=r, column=3).number_format = "#,##0"
    conv.cell(row=r, column=4).number_format = "0.00%"
    conv.cell(row=r, column=5).number_format = "#,##0"
apply_body_font(conv, len(conv_headers))
autosize(conv, [10, 28, 14, 16, 14, 50])
conv.freeze_panes = "B2"


# ===== Sheet 5 — Dashboard ==================================================

dash = wb.create_sheet("Dashboard")

# Layout: three zones stacked. Plain labels in col A, formulas in col B (and C
# where needed). All Arial, no fills, headers bold.

def set_label(cell, value, *, bold=False):
    cell.value = value
    cell.font = Font(name=ARIAL, bold=bold)


# ---- Zone 1: Overall Health -----------------------------------------------
set_label(dash.cell(row=1, column=1), "Overall Health", bold=True)

set_label(dash.cell(row=2, column=1), "Total videos logged")
dash.cell(row=2, column=2).value = (
    f"=COUNTA(Identity!C2:C{ROWS + 1})"
)  # count by Final Title

set_label(dash.cell(row=3, column=1), "Total 7-day views")
dash.cell(row=3, column=2).value = f"=SUM(Performance!H2:H{ROWS + 1})"
dash.cell(row=3, column=2).number_format = "#,##0"

set_label(dash.cell(row=4, column=1), "Total 30-day views")
dash.cell(row=4, column=2).value = f"=SUM(Performance!I2:I{ROWS + 1})"
dash.cell(row=4, column=2).number_format = "#,##0"

set_label(dash.cell(row=5, column=1), "Avg 24hr CTR%")
dash.cell(row=5, column=2).value = (
    f"=IFERROR(AVERAGE(Performance!B2:B{ROWS + 1}),0)"
)
dash.cell(row=5, column=2).number_format = "0.00%"

set_label(dash.cell(row=6, column=1), "Avg 7-day CTR%")
dash.cell(row=6, column=2).value = (
    f"=IFERROR(AVERAGE(Performance!C2:C{ROWS + 1}),0)"
)
dash.cell(row=6, column=2).number_format = "0.00%"

set_label(dash.cell(row=7, column=1), "Avg 30-day CTR%")
dash.cell(row=7, column=2).value = (
    f"=IFERROR(AVERAGE(Performance!D2:D{ROWS + 1}),0)"
)
dash.cell(row=7, column=2).number_format = "0.00%"

set_label(dash.cell(row=8, column=1), "Avg drop-off rate%")
dash.cell(row=8, column=2).value = (
    f"=IFERROR(AVERAGE(Performance!E2:E{ROWS + 1}),0)"
)
dash.cell(row=8, column=2).number_format = "0.00%"

set_label(dash.cell(row=9, column=1), "Total calls booked")
dash.cell(row=9, column=2).value = f"=SUM(Conversion!E2:E{ROWS + 1})"

# Best / worst by 7-day CTR — looked up by Final Title via INDEX/MATCH.
set_label(dash.cell(row=10, column=1), "Best video (7-day CTR)")
dash.cell(row=10, column=2).value = (
    f"=IFERROR(INDEX(Identity!C2:C{ROWS + 1},"
    f"MATCH(MAX(Performance!C2:C{ROWS + 1}),Performance!C2:C{ROWS + 1},0)),\"\")"
)
set_label(dash.cell(row=10, column=3), "7-day CTR")
dash.cell(row=10, column=4).value = f"=IFERROR(MAX(Performance!C2:C{ROWS + 1}),0)"
dash.cell(row=10, column=4).number_format = "0.00%"

set_label(dash.cell(row=11, column=1), "Worst video (7-day CTR)")
dash.cell(row=11, column=2).value = (
    f"=IFERROR(INDEX(Identity!C2:C{ROWS + 1},"
    f"MATCH(MIN(Performance!C2:C{ROWS + 1}),Performance!C2:C{ROWS + 1},0)),\"\")"
)
set_label(dash.cell(row=11, column=3), "7-day CTR")
dash.cell(row=11, column=4).value = f"=IFERROR(MIN(Performance!C2:C{ROWS + 1}),0)"
dash.cell(row=11, column=4).number_format = "0.00%"


# ---- Zone 2: Pattern Recognition -------------------------------------------
set_label(dash.cell(row=14, column=1), "Pattern Recognition", bold=True)

# By Winning Style (Creative col F)
set_label(dash.cell(row=15, column=1), "Avg 7-day CTR by Title Style", bold=True)
title_styles = ["Curiosity", "List", "How-To", "Question", "Shock"]
for i, style in enumerate(title_styles):
    r = 16 + i
    set_label(dash.cell(row=r, column=1), style)
    dash.cell(row=r, column=2).value = (
        f"=IFERROR(AVERAGEIFS(Performance!C2:C{ROWS + 1},"
        f"Creative!F2:F{ROWS + 1},\"{style}\"),0)"
    )
    dash.cell(row=r, column=2).number_format = "0.00%"

# By Hook Style (Creative col O)
set_label(dash.cell(row=22, column=1), "Avg 7-day CTR by Hook Style", bold=True)
hook_styles = ["Question", "Stat", "Story", "Bold Claim", "Pattern Interrupt"]
for i, style in enumerate(hook_styles):
    r = 23 + i
    set_label(dash.cell(row=r, column=1), style)
    dash.cell(row=r, column=2).value = (
        f"=IFERROR(AVERAGEIFS(Performance!C2:C{ROWS + 1},"
        f"Creative!O2:O{ROWS + 1},\"{style}\"),0)"
    )
    dash.cell(row=r, column=2).number_format = "0.00%"

# By Face Y/N (Creative col H)
set_label(dash.cell(row=29, column=1), "Avg 7-day CTR by Face on Thumbnail", bold=True)
for i, val in enumerate(["Y", "N"]):
    r = 30 + i
    set_label(dash.cell(row=r, column=1), f"Face = {val}")
    dash.cell(row=r, column=2).value = (
        f"=IFERROR(AVERAGEIFS(Performance!C2:C{ROWS + 1},"
        f"Creative!H2:H{ROWS + 1},\"{val}\"),0)"
    )
    dash.cell(row=r, column=2).number_format = "0.00%"

# By Word Count bucket (Creative col J)
set_label(dash.cell(row=33, column=1), "Avg 7-day CTR by Word Count Bucket", bold=True)
buckets = [
    ("1-3 words", 1, 3),
    ("4-6 words", 4, 6),
    ("7-9 words", 7, 9),
    ("10+ words", 10, 999),
]
for i, (label, lo, hi) in enumerate(buckets):
    r = 34 + i
    set_label(dash.cell(row=r, column=1), label)
    dash.cell(row=r, column=2).value = (
        f"=IFERROR(AVERAGEIFS(Performance!C2:C{ROWS + 1},"
        f"Creative!J2:J{ROWS + 1},\">={lo}\","
        f"Creative!J2:J{ROWS + 1},\"<={hi}\"),0)"
    )
    dash.cell(row=r, column=2).number_format = "0.00%"

# Correlation: intro length (Creative col P) vs drop-off rate (Performance col E)
set_label(dash.cell(row=39, column=1), "Correlation: Intro Length vs Drop-off")
dash.cell(row=39, column=2).value = (
    f"=IFERROR(CORREL(Creative!P2:P{ROWS + 1},Performance!E2:E{ROWS + 1}),\"n/a\")"
)
dash.cell(row=39, column=2).number_format = "0.000"


# ---- Zone 3: Top performers -----------------------------------------------
set_label(dash.cell(row=42, column=1), "Top 5 by 7-day CTR", bold=True)
top_headers = [
    "Rank",
    "Video ID",
    "Final Title",
    "7-day CTR",
    "7-day Views",
    "Drop-off Rate",
    "Winning Style",
    "Hook Style",
    "Calls Booked",
]
for col_idx, label in enumerate(top_headers, start=1):
    c = dash.cell(row=43, column=col_idx, value=label)
    c.font = HEADER_FONT

# LARGE() picks the nth-highest value; INDEX/MATCH grabs context for that row.
# Note: ties are not deterministic — good enough for this tracker.
for n in range(1, 6):
    r = 43 + n
    set_label(dash.cell(row=r, column=1), n)
    ctr_lookup = (
        f"LARGE(Performance!C2:C{ROWS + 1},{n})"
    )
    match_row = (
        f"MATCH({ctr_lookup},Performance!C2:C{ROWS + 1},0)"
    )
    # Video ID
    dash.cell(row=r, column=2).value = (
        f"=IFERROR(INDEX(Identity!A2:A{ROWS + 1},{match_row}),\"\")"
    )
    # Final Title
    dash.cell(row=r, column=3).value = (
        f"=IFERROR(INDEX(Identity!C2:C{ROWS + 1},{match_row}),\"\")"
    )
    # 7-day CTR
    dash.cell(row=r, column=4).value = f"=IFERROR({ctr_lookup},\"\")"
    dash.cell(row=r, column=4).number_format = "0.00%"
    # 7-day Views
    dash.cell(row=r, column=5).value = (
        f"=IFERROR(INDEX(Performance!H2:H{ROWS + 1},{match_row}),\"\")"
    )
    dash.cell(row=r, column=5).number_format = "#,##0"
    # Drop-off Rate
    dash.cell(row=r, column=6).value = (
        f"=IFERROR(INDEX(Performance!E2:E{ROWS + 1},{match_row}),\"\")"
    )
    dash.cell(row=r, column=6).number_format = "0.00%"
    # Winning Style
    dash.cell(row=r, column=7).value = (
        f"=IFERROR(INDEX(Creative!F2:F{ROWS + 1},{match_row}),\"\")"
    )
    # Hook Style
    dash.cell(row=r, column=8).value = (
        f"=IFERROR(INDEX(Creative!O2:O{ROWS + 1},{match_row}),\"\")"
    )
    # Calls Booked
    dash.cell(row=r, column=9).value = (
        f"=IFERROR(INDEX(Conversion!E2:E{ROWS + 1},{match_row}),\"\")"
    )

# Apply Arial body font to all data cells in dashboard.
for row in dash.iter_rows(min_row=1, max_row=50, min_col=1, max_col=9):
    for c in row:
        if c.font.bold:
            c.font = Font(name=ARIAL, bold=True)
        else:
            c.font = BODY_FONT

autosize(dash, [34, 28, 18, 14, 14, 16, 16, 18, 14])


# ===== Sheet 6 — AI Export ==================================================

ai = wb.create_sheet("AI Export")

# A1: Context blurb describing the schema for downstream Gemini prompts.
context_text = (
    "CONTEXT FOR AI: This sheet is a flat denormalized export of a YouTube "
    "content tracker. Each row (starting row 3) is one video, joined across "
    "four source sheets (Identity, Creative, Performance, Conversion) by "
    "Video ID. Columns in row 2 are headers. Identity captures publishing "
    "metadata. Creative captures title/thumbnail/hook design choices and "
    "A/B variants. Performance captures click-through and retention metrics "
    "at 24hr / 7-day / 30-day windows plus drop-off and average view "
    "duration. Conversion captures CTAs, click-throughs, engagement rate, "
    "calls booked, and free-form notes. CTR / drop-off / engagement values "
    "are stored as decimals (e.g., 0.05 = 5%). Use this view to identify "
    "which creative levers (title style, hook style, face on thumbnail, "
    "word count, intro length) correlate with higher CTR, lower drop-off, "
    "and more calls booked. Empty cells mean no data was logged."
)
ai.cell(row=1, column=1, value=context_text).font = BODY_FONT
ai.cell(row=1, column=1).alignment = Alignment(
    wrap_text=True, vertical="top", horizontal="left"
)
ai.row_dimensions[1].height = 90

# Row 2: full denormalized header set, ordered Identity → Creative →
# Performance → Conversion. Skip duplicate Video ID columns from sister
# sheets — Video ID appears once at column A.
ai_headers: list[tuple[str, str, str]] = []  # (header label, source sheet, source col)
# Identity (cols A-F)
for letter, label in zip("ABCDEF", identity_headers):
    ai_headers.append((label, "Identity", letter))
# Creative (skip A = Video ID; cols B-Q)
creative_letters = [get_column_letter(i) for i in range(2, len(creative_headers) + 1)]
for letter, label in zip(creative_letters, creative_headers[1:]):
    ai_headers.append((label, "Creative", letter))
# Performance (skip A; cols B-I)
perf_letters = [get_column_letter(i) for i in range(2, len(perf_headers) + 1)]
for letter, label in zip(perf_letters, perf_headers[1:]):
    ai_headers.append((label, "Performance", letter))
# Conversion (skip A; cols B-F)
conv_letters = [get_column_letter(i) for i in range(2, len(conv_headers) + 1)]
for letter, label in zip(conv_letters, conv_headers[1:]):
    ai_headers.append((label, "Conversion", letter))

for col_idx, (label, _src, _letter) in enumerate(ai_headers, start=1):
    c = ai.cell(row=2, column=col_idx, value=label)
    c.font = HEADER_FONT

# Rows 3+: pull each cell via INDEX/MATCH on Video ID. Identity columns are a
# direct passthrough (already keyed by Video ID); the others use MATCH against
# Identity!A.
for i in range(ROWS):
    out_row = i + 3
    src_row = i + 2  # corresponding Identity row
    for col_idx, (_label, src, letter) in enumerate(ai_headers, start=1):
        cell = ai.cell(row=out_row, column=col_idx)
        if src == "Identity":
            cell.value = f"=IF(Identity!{letter}{src_row}=\"\",\"\",Identity!{letter}{src_row})"
        else:
            # INDEX/MATCH join on Video ID.
            cell.value = (
                f"=IFERROR(INDEX({src}!{letter}:{letter},"
                f"MATCH(Identity!A{src_row},{src}!A:A,0)),\"\")"
            )
        cell.font = BODY_FONT

# Reasonable default widths.
ai_widths = [14] * len(ai_headers)
autosize(ai, ai_widths)
ai.freeze_panes = "B3"


# ---- save ------------------------------------------------------------------

import os

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tracker.xlsx")
wb.save(out_path)
print(f"Wrote {out_path}")
